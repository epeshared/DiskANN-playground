#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import html
import io
import json
import os
import re
import math
import shutil
import subprocess
import urllib.parse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.requests import Request
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from markdown_it import MarkdownIt


@dataclass(frozen=True)
class Settings:
    host: str
    port: int
    runs_dir: Path


def _env(name: str, default: str) -> str:
    v = os.environ.get(name)
    return v if v is not None and v.strip() else default


def _safe_name(s: str) -> str:
    # Keep URLs simple and prevent traversal. Dataset and run_id are directory names.
    if not re.fullmatch(r"[A-Za-z0-9._-]+", s or ""):
        raise HTTPException(status_code=400, detail=f"Invalid path component: {s!r}")
    return s


def _safe_join(base: Path, *parts: str) -> Path:
    p = base
    for part in parts:
        p = p / _safe_name(part)
    rp = p.resolve()
    rb = base.resolve()
    if rb != rp and rb not in rp.parents:
        raise HTTPException(status_code=400, detail="Path traversal blocked")
    return rp


def _list_dirs(path: Path) -> list[str]:
    if not path.exists():
        return []
    out: list[str] = []
    for child in path.iterdir():
        if child.is_dir():
            out.append(child.name)
    out.sort()
    return out


def _list_dirs_sorted_desc(path: Path) -> list[str]:
    xs = _list_dirs(path)
    xs.sort(reverse=True)
    return xs


def _read_text_if_exists(path: Path, *, max_bytes: int = 2_000_000) -> str | None:
    try:
        if not path.is_file():
            return None
        if path.stat().st_size > max_bytes:
            return path.read_text(encoding="utf-8", errors="replace")[:max_bytes] + "\n<truncated>\n"
        return path.read_text(encoding="utf-8", errors="replace")
    except Exception:
        return None


def _read_tsv(path: Path) -> tuple[list[str], list[list[str]]]:
    if not path.is_file():
        return ([], [])
    raw = path.read_text(encoding="utf-8", errors="replace").splitlines()
    if not raw:
        return ([], [])
    headers = raw[0].split("\t")
    rows = [line.split("\t") for line in raw[1:] if line.strip()]
    return (headers, rows)


def _read_csv(path: Path) -> tuple[list[str], list[list[str]]]:
    if not path.is_file():
        return ([], [])
    try:
        import csv

        with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
            reader = csv.reader(f)
            raw = list(reader)
    except Exception:
        return ([], [])

    if not raw:
        return ([], [])
    headers = [c.strip() for c in raw[0]]
    rows = [row for row in raw[1:] if any((c or "").strip() for c in row)]
    return (headers, rows)


def _to_float(v: str | None) -> float | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        x = float(s)
    except Exception:
        return None
    if not math.isfinite(x):
        return None
    return x


def _mean_csv_col(path: Path, col_name: str) -> float | None:
    headers, rows = _read_csv(path)
    if not headers or not rows:
        return None
    try:
        idx = headers.index(col_name)
    except ValueError:
        return None

    vals: list[float] = []
    for row in rows:
        if idx >= len(row):
            continue
        x = _to_float(row[idx])
        if x is not None:
            vals.append(x)
    if not vals:
        return None
    return sum(vals) / float(len(vals))


def _fmt_float(v: float | None, digits: int = 2) -> str:
    if v is None:
        return ""
    return f"{v:.{digits}f}"


def _read_csv_dict_rows(path: Path) -> list[dict[str, str]]:
    headers, rows = _read_csv(path)
    if not headers or not rows:
        return []
    out: list[dict[str, str]] = []
    for row in rows:
        if len(row) < len(headers):
            row = row + [""] * (len(headers) - len(row))
        out.append({h: (row[i] if i < len(row) else "") for i, h in enumerate(headers)})
    return out


def _compare_csv_by_keys(path_a: Path, path_b: Path, key_fields: list[str]) -> dict[str, Any]:
    rows_a = _read_csv_dict_rows(path_a)
    rows_b = _read_csv_dict_rows(path_b)
    if not rows_a or not rows_b:
        return {
            "available": False,
            "key_fields": key_fields,
            "rows": [],
        }

    def make_key(row: dict[str, str]) -> tuple[str, ...]:
        return tuple((row.get(k) or "").strip() for k in key_fields)

    map_a: dict[tuple[str, ...], dict[str, str]] = {}
    map_b: dict[tuple[str, ...], dict[str, str]] = {}

    for r in rows_a:
        map_a[make_key(r)] = r
    for r in rows_b:
        map_b[make_key(r)] = r

    all_keys = sorted(set(map_a.keys()) | set(map_b.keys()))
    out_rows: list[dict[str, str]] = []

    def _ratio_str(a: float | None, b: float | None) -> str:
        if a is None or b is None or b == 0.0:
            return ""
        return _fmt_float(a / b, 3)

    for key in all_keys:
        ra = map_a.get(key, {})
        rb = map_b.get(key, {})
        row: dict[str, str] = {}
        for i, k in enumerate(key_fields):
            row[k] = key[i]

        recall_a_v = _to_float(ra.get("recall_at_k"))
        recall_b_v = _to_float(rb.get("recall_at_k"))
        qps_a_v = _to_float(ra.get("qps_mean"))
        qps_b_v = _to_float(rb.get("qps_mean"))
        p99_a_v = _to_float(ra.get("lat_p99_us"))
        p99_b_v = _to_float(rb.get("lat_p99_us"))

        row["recall_a"] = _fmt_float(recall_a_v)
        row["recall_b"] = _fmt_float(recall_b_v)
        row["recall_ratio"] = _ratio_str(recall_a_v, recall_b_v)
        row["qps_a"] = _fmt_float(qps_a_v)
        row["qps_b"] = _fmt_float(qps_b_v)
        row["qps_ratio"] = _ratio_str(qps_a_v, qps_b_v)
        row["p99_a"] = _fmt_float(p99_a_v)
        row["p99_b"] = _fmt_float(p99_b_v)
        row["p99_ratio"] = _ratio_str(p99_a_v, p99_b_v)

        out_rows.append(row)

    return {
        "available": True,
        "key_fields": key_fields,
        "rows": out_rows,
    }


def _reorder_table(
    headers: list[str],
    rows: list[list[str]],
    *,
    left_preferred: list[str],
    right_preferred: list[str],
) -> tuple[list[str], list[list[str]]]:
    if not headers:
        return (headers, rows)

    header_set = set(headers)
    left = [h for h in left_preferred if h in header_set]
    right = [h for h in right_preferred if h in header_set]

    # Middle columns keep the original order.
    middle = [h for h in headers if h not in set(left) and h not in set(right)]
    new_headers = left + middle + right

    idx = {h: i for i, h in enumerate(headers)}
    new_rows: list[list[str]] = []
    for r in rows:
        if len(r) < len(headers):
            r = r + [""] * (len(headers) - len(r))
        new_rows.append([r[idx[h]] if idx[h] < len(r) else "" for h in new_headers])

    return (new_headers, new_rows)


def _drop_columns(headers: list[str], rows: list[list[str]], *, drop: set[str]) -> tuple[list[str], list[list[str]]]:
    if not headers or not drop:
        return (headers, rows)

    keep_indices = [i for i, h in enumerate(headers) if h not in drop]
    new_headers = [headers[i] for i in keep_indices]

    new_rows: list[list[str]] = []
    for r in rows:
        if len(r) < len(headers):
            r = r + [""] * (len(headers) - len(r))
        new_rows.append([r[i] if i < len(r) else "" for i in keep_indices])

    return (new_headers, new_rows)


def _parse_cpu_bind_to_set(cpu_bind: str) -> set[int]:
    s = (cpu_bind or '').strip()
    if not s:
        return set()

    # Accept:
    # - "0-16"
    # - "0,1,2,3"
    # - "0 1 2"
    # - "physcpubind: 0 1 2"
    s = s.replace('physcpubind:', ' ')
    toks = re.split(r"[\s,]+", s)
    out: set[int] = set()
    for t in toks:
        t = t.strip()
        if not t:
            continue
        if '-' in t:
            a, b = t.split('-', 1)
            if a.strip().isdigit() and b.strip().isdigit():
                lo = int(a)
                hi = int(b)
                if lo > hi:
                    lo, hi = hi, lo
                for x in range(lo, hi + 1):
                    out.add(x)
                continue
        if t.isdigit():
            out.add(int(t))
    return out


def _breadcrumbs(*items: tuple[str, str]) -> str:
    parts: list[str] = []
    for label, href in items:
        parts.append(f'<a href="{html.escape(href, quote=True)}">{html.escape(label)}</a>')
    return " / ".join(parts)


def create_app(settings: Settings) -> FastAPI:
    app = FastAPI(title="diskann-ann-bench web")

    templates_dir = Path(__file__).resolve().parent / "templates"
    templates = Jinja2Templates(directory=str(templates_dir))
    templates.env.filters["urlencode"] = lambda s: urllib.parse.quote(str(s), safe="")

    static_dir = Path(__file__).resolve().parent / "static"
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    md = MarkdownIt("commonmark")

    # Best-effort server info cache (evaluated lazily).
    lscpu_cache: dict[str, str | None] = {"raw": None}

    def _read_lscpu_raw() -> str | None:
        try:
            out = subprocess.check_output(["lscpu"], stderr=subprocess.STDOUT, text=True, timeout=2.0)
            return out
        except Exception:
            return None

    def _lscpu_summary(raw: str) -> str:
        lines = raw.splitlines()

        def find_first(prefix: str) -> str | None:
            for ln in lines:
                if ln.lstrip().startswith(prefix):
                    return ln.rstrip()
            return None

        out: list[str] = []

        # Match the user-requested sections/order as closely as possible.
        for key in [
            "Architecture:",
            "CPU op-mode(s):",
            "Address sizes:",
            "Byte Order:",
            "CPU(s):",
            "On-line CPU(s) list:",
            "Vendor ID:",
            "Model name:",
            "CPU family:",
            "Model:",
            "Thread(s) per core:",
            "Core(s) per socket:",
            "Socket(s):",
            "Stepping:",
            "CPU max MHz:",
            "CPU min MHz:",
            "BogoMIPS:",
        ]:
            v = find_first(key)
            if v:
                out.append(v)

        # Virtualization features
        vf_hdr = find_first("Virtualization features:")
        if vf_hdr:
            out.append(vf_hdr)
        vtx = find_first("Virtualization:")
        if vtx:
            out.append(vtx)

        # Cache summary
        caches_hdr = find_first("Caches (sum of all):")
        if caches_hdr:
            out.append(caches_hdr)
        for key in ["L1d:", "L1i:", "L2:", "L3:"]:
            v = find_first(key)
            if v:
                out.append(v)

        # NUMA
        numa_hdr = find_first("NUMA:")
        if numa_hdr:
            out.append(numa_hdr)
        nn = find_first("NUMA node(s):")
        if nn:
            out.append(nn)
        # Include all NUMA node CPU lines if present.
        for ln in lines:
            if re.match(r"^\s*NUMA node\d+ CPU\(s\):", ln):
                out.append(ln.rstrip())

        return "\n".join(out).strip()

    def _lscpu_model_name(raw: str) -> str | None:
        for ln in raw.splitlines():
            s = ln.lstrip()
            if s.startswith("Model name:"):
                parts = s.split(":", 1)
                if len(parts) == 2:
                    v = parts[1].strip()
                    return v or None
                return None
        return None

    def _compare_algo_rows(run_a_dir: Path, run_b_dir: Path, algo: str) -> dict[str, Any]:
        algo_norm = (algo or "").strip().lower()
        if algo_norm == "pq":
            return _compare_csv_by_keys(
                run_a_dir / "outputs" / "summary.pq.csv",
                run_b_dir / "outputs" / "summary.pq.csv",
                [
                    "case",
                    "distance",
                    "k",
                    "l_search",
                    "reps",
                    "l_build",
                    "max_outdegree",
                    "alpha",
                    "num_pq_chunks",
                    "translate_to_center",
                    "num_centers",
                    "max_k_means_reps",
                    "pq_seed",
                ],
            )
        if algo_norm == "spherical":
            return _compare_csv_by_keys(
                run_a_dir / "outputs" / "summary.spherical.csv",
                run_b_dir / "outputs" / "summary.spherical.csv",
                [
                    "case",
                    "distance",
                    "k",
                    "l_search",
                    "reps",
                    "l_build",
                    "max_outdegree",
                    "alpha",
                    "nbits",
                    "spherical_seed",
                ],
            )
        raise HTTPException(status_code=400, detail="algo must be one of: pq, spherical")

    def _run_meta(run_dir: Path, run_id: str) -> dict[str, Any]:
        mode_txt = _read_text_if_exists(run_dir / "mode.txt", max_bytes=2000)
        run_mode = (mode_txt or "").strip() or None

        batch_txt = _read_text_if_exists(run_dir / "batch.txt", max_bytes=50)
        batch_norm = (batch_txt or "").strip().lower()
        is_batch = batch_norm in {"1", "true", "yes", "y", "batch"}
        query_mode = "batch" if is_batch else "single"

        cpu_bind_txt = _read_text_if_exists(run_dir / "cpu-bind.txt", max_bytes=5000)
        cpu_bind = (cpu_bind_txt or "").strip() or None

        lscpu_raw = _read_text_if_exists(run_dir / "lscpu.txt", max_bytes=500_000)
        if not lscpu_raw:
            if lscpu_cache.get("raw") is None:
                lscpu_cache["raw"] = _read_lscpu_raw()
            lscpu_raw = lscpu_cache.get("raw")

        cpu_model = _lscpu_model_name(lscpu_raw) if isinstance(lscpu_raw, str) and lscpu_raw.strip() else None
        lscpu_info = _lscpu_summary(lscpu_raw) if isinstance(lscpu_raw, str) and lscpu_raw.strip() else None

        return {
            "id": run_id,
            "mode": run_mode,
            "query_mode": query_mode,
            "cpu_bind": cpu_bind,
            "cpu_model": cpu_model,
            "lscpu_info": lscpu_info,
        }

    default_mode = "ann_bench_diskann_rs"

    @app.get("/api/health")
    def health() -> dict[str, Any]:
        return {"ok": True, "runs_dir": str(settings.runs_dir)}

    @app.get("/api/datasets")
    def api_datasets() -> JSONResponse:
        return JSONResponse(_list_dirs(settings.runs_dir))

    @app.get("/api/runs/{dataset}")
    def api_runs(dataset: str) -> JSONResponse:
        dataset_dir = _safe_join(settings.runs_dir, dataset)
        return JSONResponse(_list_dirs_sorted_desc(dataset_dir))

    @app.get("/api/file/{dataset}/{run_id}/{rel_path:path}")
    def api_file(dataset: str, run_id: str, rel_path: str):
        # Lock downloads to files under the run directory.
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        # Disallow weird components in rel_path.
        rel_path = rel_path.lstrip("/")
        if any(x in rel_path for x in ["..", "\\", "//"]):
            raise HTTPException(status_code=400, detail="Invalid rel_path")
        path = (run_dir / rel_path).resolve()
        if run_dir != path and run_dir not in path.parents:
            raise HTTPException(status_code=400, detail="Path traversal blocked")
        if not path.is_file():
            raise HTTPException(status_code=404, detail="File not found")
        return FileResponse(str(path))

    @app.get("/", response_class=HTMLResponse)
    def index(request: Request):
        datasets = _list_dirs(settings.runs_dir)
        return templates.TemplateResponse(
            "index.html",
            {
                "request": request,
                "title": "diskann-ann-bench",
                "datasets": datasets,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": None,
            },
        )

    @app.get("/dataset/{dataset}", response_class=HTMLResponse)
    def dataset_page(dataset: str, request: Request):
        dataset_dir = _safe_join(settings.runs_dir, dataset)
        run_ids = _list_dirs_sorted_desc(dataset_dir)

        q = (request.query_params.get("q") or "").strip().lower() or None
        mode = request.query_params.get("mode")
        mode = (mode.strip() if isinstance(mode, str) else "")
        mode_norm = (mode.lower().strip() if mode else "") or default_mode
        cpu_filter = request.query_params.get("cpu")
        cpu_filter = (cpu_filter.strip() if isinstance(cpu_filter, str) else "")
        cpu_filter_norm = cpu_filter.lower()
        cpu_bind_filter = request.query_params.get("cpu_bind")
        cpu_bind_filter = (cpu_bind_filter.strip() if isinstance(cpu_bind_filter, str) else "")
        cpu_bind_filter_norm = cpu_bind_filter.lower()

        batch_filter = request.query_params.get("batch")
        batch_filter = (batch_filter.strip() if isinstance(batch_filter, str) else "")
        batch_filter_norm = batch_filter.lower().strip()
        deleted = request.query_params.get("deleted")
        try:
            deleted_count = int(deleted) if deleted is not None else 0
        except Exception:
            deleted_count = 0

        runs = []
        for run_id in run_ids:
            run_dir = dataset_dir / run_id
            mode_txt = _read_text_if_exists(run_dir / "mode.txt", max_bytes=2000)
            run_mode = (mode_txt or "").strip() or None

            lscpu_raw = _read_text_if_exists(run_dir / "lscpu.txt", max_bytes=200_000)
            if not lscpu_raw:
                if lscpu_cache.get("raw") is None:
                    lscpu_cache["raw"] = _read_lscpu_raw()
                lscpu_raw = lscpu_cache.get("raw")
            cpu_model = _lscpu_model_name(lscpu_raw) if isinstance(lscpu_raw, str) and lscpu_raw.strip() else None

            cpu_bind_txt = _read_text_if_exists(run_dir / "cpu-bind.txt", max_bytes=5000)
            cpu_bind = (cpu_bind_txt or '').strip() or None
            cpu_cores = len(_parse_cpu_bind_to_set(cpu_bind or '')) if cpu_bind else None

            memory_txt = _read_text_if_exists(run_dir / "memory.txt", max_bytes=2000)
            memory = (memory_txt or "").strip() or None

            batch_txt = _read_text_if_exists(run_dir / "batch.txt", max_bytes=50)
            batch_norm = (batch_txt or "").strip().lower()
            is_batch = batch_norm in {"1", "true", "yes", "y", "batch"}
            query_mode = "batch" if is_batch else "single"

            if mode_norm and (run_mode or "").strip().lower() != mode_norm:
                continue
            if cpu_filter_norm and cpu_filter_norm not in (cpu_model or "").lower():
                continue
            if cpu_bind_filter_norm and cpu_bind_filter_norm not in (cpu_bind or "").lower():
                continue
            if batch_filter_norm:
                if batch_filter_norm in {"1", "true", "yes", "y", "batch"} and not is_batch:
                    continue
                if batch_filter_norm in {"0", "false", "no", "n", "single"} and is_batch:
                    continue

            if q:
                hay = " ".join([run_id, run_mode or ""]).lower()
                if q not in hay:
                    continue

            runs.append(
                {
                    "id": run_id,
                    "mode": run_mode,
                    "query_mode": query_mode,
                    "cpu_model": cpu_model,
                    "memory": memory,
                    "cpu_bind": cpu_bind,
                    "cpu_cores": cpu_cores,
                    "has_summary": (
                        (run_dir / "outputs" / "summary.pq.csv").is_file()
                        or (run_dir / "outputs" / "summary.spherical.csv").is_file()
                        or (run_dir / "outputs" / "summary.tsv").is_file()
                    ),
                    "has_details": (run_dir / "outputs" / "details.md").is_file(),
                    "has_build": (run_dir / "outputs" / "output.build.json").is_file(),
                    "has_search": (run_dir / "outputs" / "output.search.json").is_file(),
                }
            )

        return templates.TemplateResponse(
            "dataset.html",
            {
                "request": request,
                "title": f"{dataset}",
                "dataset": dataset,
                "runs": runs,
                "q": q or "",
                "mode": mode,
                "cpu_filter": cpu_filter,
                "cpu_bind_filter": cpu_bind_filter,
                "batch_filter": batch_filter,
                "default_mode": default_mode,
                "deleted_count": deleted_count,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": _breadcrumbs(("datasets", "/"), (dataset, f"/dataset/{urllib.parse.quote(dataset, safe='')}")),
            },
        )

    @app.post("/dataset/{dataset}/delete")
    async def dataset_delete_runs(dataset: str, request: Request):
        dataset_dir = _safe_join(settings.runs_dir, dataset)
        form = await request.form()

        selected: list[str] = []
        if hasattr(form, "getlist"):
            selected = [str(x).strip() for x in form.getlist("run_ids") if str(x).strip()]

        q = (str(form.get("q") or "")).strip()
        mode = (str(form.get("mode") or "")).strip()
        cpu_filter = (str(form.get("cpu") or "")).strip()
        cpu_bind_filter = (str(form.get("cpu_bind") or "")).strip()
        batch_filter = (str(form.get("batch") or "")).strip()

        deleted_count = 0
        for run_id in selected:
            try:
                run_dir = _safe_join(dataset_dir, run_id)
            except HTTPException:
                continue
            if run_dir.is_dir():
                shutil.rmtree(run_dir)
                deleted_count += 1

        base = f"/dataset/{urllib.parse.quote(dataset, safe='')}"
        params: dict[str, str] = {"deleted": str(deleted_count)}
        if q:
            params["q"] = q
        if mode:
            params["mode"] = mode
        if cpu_filter:
            params["cpu"] = cpu_filter
        if cpu_bind_filter:
            params["cpu_bind"] = cpu_bind_filter
        if batch_filter:
            params["batch"] = batch_filter
        url = f"{base}?{urllib.parse.urlencode(params)}"
        return RedirectResponse(url=url, status_code=303)

    @app.get("/run/{dataset}/{run_id}", response_class=HTMLResponse)
    def run_page(dataset: str, run_id: str, request: Request):
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)

        mode_txt = _read_text_if_exists(run_dir / "mode.txt", max_bytes=2000)
        run_mode = (mode_txt or "").strip() or None

        batch_txt = _read_text_if_exists(run_dir / "batch.txt", max_bytes=50)
        batch_norm = (batch_txt or "").strip().lower()
        is_batch = batch_norm in {"1", "true", "yes", "y", "batch"}
        query_mode = "batch" if is_batch else "single"

        cpu_bind_txt = _read_text_if_exists(run_dir / "cpu-bind.txt", max_bytes=5000)
        cpu_bind = (cpu_bind_txt or '').strip() or None
        cpu_cores = len(_parse_cpu_bind_to_set(cpu_bind or '')) if cpu_bind else None

        pq_headers, pq_rows = _read_csv(run_dir / "outputs" / "summary.pq.csv")
        spherical_headers, spherical_rows = _read_csv(run_dir / "outputs" / "summary.spherical.csv")
        headers, rows = _read_tsv(run_dir / "outputs" / "summary.tsv")

        perf_right = [
            "peak_rss_gib",
            "peak_tps",
            "peak_kB_read_s",
            "peak_kB_wrtn_s",
            "recall_at_k",
            "qps_mean",
            "lat_mean_us",
            "lat_p50_us",
            "lat_p95_us",
            "lat_p99_us",
            "load_index_s",
        ]

        pq_left = [
            "case",
            "distance",
            "k",
            "l_search",
            "reps",
            "l_build",
            "max_outdegree",
            "alpha",
            "num_pq_chunks",
            "translate_to_center",
            "num_centers",
            "max_k_means_reps",
            "pq_seed",
            "index_prefix",
        ]
        pq_headers, pq_rows = _reorder_table(
            pq_headers,
            pq_rows,
            left_preferred=pq_left,
            right_preferred=perf_right,
        )
        pq_headers, pq_rows = _drop_columns(pq_headers, pq_rows, drop={"index_prefix"})

        spherical_left = [
            "case",
            "distance",
            "k",
            "l_search",
            "reps",
            "l_build",
            "max_outdegree",
            "alpha",
            "nbits",
            "spherical_seed",
            "index_prefix",
        ]
        spherical_headers, spherical_rows = _reorder_table(
            spherical_headers,
            spherical_rows,
            left_preferred=spherical_left,
            right_preferred=perf_right,
        )
        spherical_headers, spherical_rows = _drop_columns(spherical_headers, spherical_rows, drop={"index_prefix"})

        # Best-effort ordering for legacy TSV.
        legacy_left = [
            "case",
            "algo",
            "distance",
            "k",
            "L",
            "l_search",
            "reps",
            "l_build",
            "max_outdegree",
            "alpha",
            "index_prefix",
        ]
        headers, rows = _reorder_table(
            headers,
            rows,
            left_preferred=legacy_left,
            right_preferred=perf_right,
        )
        headers, rows = _drop_columns(headers, rows, drop={"index_prefix"})
        details_md = _read_text_if_exists(run_dir / "outputs" / "details.md")
        details_html = md.render(details_md) if details_md else None

        lscpu_raw = _read_text_if_exists(run_dir / "lscpu.txt", max_bytes=500_000)
        if not lscpu_raw:
            if lscpu_cache.get("raw") is None:
                lscpu_cache["raw"] = _read_lscpu_raw()
            lscpu_raw = lscpu_cache.get("raw")
        lscpu_info = _lscpu_summary(lscpu_raw) if isinstance(lscpu_raw, str) and lscpu_raw.strip() else None

        build_json_obj = None
        search_json_obj = None

        build_raw = _read_text_if_exists(run_dir / "outputs" / "output.build.json")
        if build_raw:
            try:
                build_json_obj = json.dumps(json.loads(build_raw), indent=2)
            except Exception:
                build_json_obj = build_raw

        search_raw = _read_text_if_exists(run_dir / "outputs" / "output.search.json")
        if search_raw:
            try:
                search_json_obj = json.dumps(json.loads(search_raw), indent=2)
            except Exception:
                search_json_obj = search_raw

        return templates.TemplateResponse(
            "run.html",
            {
                "request": request,
                "title": f"{dataset}/{run_id}",
                "dataset": dataset,
                "run_id": run_id,
                "mode": run_mode,
                "query_mode": query_mode,
                "cpu_bind": cpu_bind,
                "cpu_cores": cpu_cores,
                "summary": {"headers": headers, "rows": rows},
                "summary_pq": {"headers": pq_headers, "rows": pq_rows},
                "summary_spherical": {"headers": spherical_headers, "rows": spherical_rows},
                "details_html": details_html,
                "lscpu_info": lscpu_info,
                "build_json": build_json_obj,
                "search_json": search_json_obj,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": _breadcrumbs(
                    ("datasets", "/"),
                    (dataset, f"/dataset/{urllib.parse.quote(dataset, safe='')}"),
                    (run_id, f"/run/{urllib.parse.quote(dataset, safe='')}/{urllib.parse.quote(run_id, safe='')}"),
                ),
            },
        )

    @app.get("/compare/{dataset}", response_class=HTMLResponse)
    def compare_page(dataset: str, request: Request):
        run_a = (request.query_params.get("a") or "").strip()
        run_b = (request.query_params.get("b") or "").strip()
        if not run_a or not run_b:
            raise HTTPException(status_code=400, detail="Missing query parameters: a and b")

        run_a_dir = _safe_join(settings.runs_dir, dataset, run_a)
        run_b_dir = _safe_join(settings.runs_dir, dataset, run_b)

        meta_a = _run_meta(run_a_dir, run_a)
        meta_b = _run_meta(run_b_dir, run_b)

        pq_cmp = _compare_algo_rows(run_a_dir, run_b_dir, "pq")
        spherical_cmp = _compare_algo_rows(run_a_dir, run_b_dir, "spherical")

        return templates.TemplateResponse(
            "compare.html",
            {
                "request": request,
                "title": f"compare {dataset}: {run_a} vs {run_b}",
                "dataset": dataset,
                "run_a": meta_a,
                "run_b": meta_b,
                "pq_cmp": pq_cmp,
                "spherical_cmp": spherical_cmp,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": _breadcrumbs(
                    ("datasets", "/"),
                    (dataset, f"/dataset/{urllib.parse.quote(dataset, safe='')}"),
                    ("compare", f"/compare/{urllib.parse.quote(dataset, safe='')}?a={urllib.parse.quote(run_a, safe='')}&b={urllib.parse.quote(run_b, safe='')}"),
                ),
            },
        )

    @app.get("/api/compare/{dataset}/{algo}.csv")
    def api_compare_csv(dataset: str, algo: str, a: str, b: str, ratio: str = "ab"):
        run_a = (a or "").strip()
        run_b = (b or "").strip()
        if not run_a or not run_b:
            raise HTTPException(status_code=400, detail="Missing query parameters: a and b")

        ratio_mode = (ratio or "ab").strip().lower()
        if ratio_mode not in {"ab", "ba"}:
            raise HTTPException(status_code=400, detail="ratio must be one of: ab, ba")

        run_a_dir = _safe_join(settings.runs_dir, dataset, run_a)
        run_b_dir = _safe_join(settings.runs_dir, dataset, run_b)
        meta_a = _run_meta(run_a_dir, run_a)
        meta_b = _run_meta(run_b_dir, run_b)

        cmp = _compare_algo_rows(run_a_dir, run_b_dir, algo)
        if not cmp.get("available"):
            raise HTTPException(status_code=404, detail="Comparison source CSV not found for both runs")

        ratio_label = "A/B" if ratio_mode == "ab" else "B/A"

        def _ratio_value(row: dict[str, str], left: str, right: str) -> str:
            lv = _to_float(row.get(left))
            rv = _to_float(row.get(right))
            if ratio_mode == "ab":
                if lv is None or rv is None or rv == 0.0:
                    return ""
                return _fmt_float(lv / rv, 3)
            if lv is None or rv is None or lv == 0.0:
                return ""
            return _fmt_float(rv / lv, 3)

        fields = [
            "run_a",
            "run_b",
            "a_query_mode",
            "b_query_mode",
            "a_cpu_bind",
            "b_cpu_bind",
            "a_cpu_info",
            "b_cpu_info",
            *(cmp.get("key_fields") or []),
            "recall_a",
            "recall_b",
            f"recall_ratio ({ratio_label})",
            "qps_a",
            "qps_b",
            f"qps_ratio ({ratio_label})",
            "p99_a",
            "p99_b",
            f"p99_ratio ({ratio_label})",
        ]

        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=fields)
        w.writeheader()
        for r in (cmp.get("rows") or []):
            out_row = {k: "" for k in fields}
            out_row.update({k: r.get(k, "") for k in (cmp.get("key_fields") or [])})
            out_row["run_a"] = run_a
            out_row["run_b"] = run_b
            out_row["a_query_mode"] = meta_a.get("query_mode") or ""
            out_row["b_query_mode"] = meta_b.get("query_mode") or ""
            out_row["a_cpu_bind"] = meta_a.get("cpu_bind") or ""
            out_row["b_cpu_bind"] = meta_b.get("cpu_bind") or ""
            out_row["a_cpu_info"] = meta_a.get("lscpu_info") or ""
            out_row["b_cpu_info"] = meta_b.get("lscpu_info") or ""

            out_row["recall_a"] = r.get("recall_a", "")
            out_row["recall_b"] = r.get("recall_b", "")
            out_row[f"recall_ratio ({ratio_label})"] = _ratio_value(r, "recall_a", "recall_b")

            out_row["qps_a"] = r.get("qps_a", "")
            out_row["qps_b"] = r.get("qps_b", "")
            out_row[f"qps_ratio ({ratio_label})"] = _ratio_value(r, "qps_a", "qps_b")

            out_row["p99_a"] = r.get("p99_a", "")
            out_row["p99_b"] = r.get("p99_b", "")
            out_row[f"p99_ratio ({ratio_label})"] = _ratio_value(r, "p99_a", "p99_b")

            w.writerow(out_row)

        filename = f"compare.{algo.lower()}.{run_a}.vs.{run_b}.csv"
        return PlainTextResponse(
            content=buf.getvalue(),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/compare/{dataset}/{algo}.xlsx")
    def api_compare_xlsx(dataset: str, algo: str, a: str, b: str, ratio: str = "ab"):
        run_a = (a or "").strip()
        run_b = (b or "").strip()
        if not run_a or not run_b:
            raise HTTPException(status_code=400, detail="Missing query parameters: a and b")

        ratio_mode = (ratio or "ab").strip().lower()
        if ratio_mode not in {"ab", "ba"}:
            raise HTTPException(status_code=400, detail="ratio must be one of: ab, ba")

        run_a_dir = _safe_join(settings.runs_dir, dataset, run_a)
        run_b_dir = _safe_join(settings.runs_dir, dataset, run_b)
        meta_a = _run_meta(run_a_dir, run_a)
        meta_b = _run_meta(run_b_dir, run_b)

        cmp = _compare_algo_rows(run_a_dir, run_b_dir, algo)
        if not cmp.get("available"):
            raise HTTPException(status_code=404, detail="Comparison source CSV not found for both runs")

        ratio_label = "A/B" if ratio_mode == "ab" else "B/A"

        def _ratio_value(row: dict[str, str], left: str, right: str) -> str:
            lv = _to_float(row.get(left))
            rv = _to_float(row.get(right))
            if ratio_mode == "ab":
                if lv is None or rv is None or rv == 0.0:
                    return ""
                return _fmt_float(lv / rv, 3)
            if lv is None or rv is None or lv == 0.0:
                return ""
            return _fmt_float(rv / lv, 3)

        key_fields = list(cmp.get("key_fields") or [])
        compare_fields = [
            *key_fields,
            "recall_a",
            "recall_b",
            f"recall_ratio ({ratio_label})",
            "qps_a",
            "qps_b",
            f"qps_ratio ({ratio_label})",
            "p99_a",
            "p99_b",
            f"p99_ratio ({ratio_label})",
        ]

        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Alignment  # type: ignore
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"openpyxl not available: {exc}")

        def _multiline_kv(s: str | None) -> str:
            raw = (s or "").strip()
            if not raw:
                return ""
            parts = [ln.strip() for ln in raw.splitlines() if ln.strip()]
            return "\n".join(parts)

        wb = Workbook()
        ws_compare = wb.active
        ws_compare.title = "compare"
        ws_compare.append(compare_fields)

        for r in (cmp.get("rows") or []):
            out = {k: "" for k in compare_fields}
            for k in key_fields:
                out[k] = r.get(k, "")
            out["recall_a"] = r.get("recall_a", "")
            out["recall_b"] = r.get("recall_b", "")
            out[f"recall_ratio ({ratio_label})"] = _ratio_value(r, "recall_a", "recall_b")
            out["qps_a"] = r.get("qps_a", "")
            out["qps_b"] = r.get("qps_b", "")
            out[f"qps_ratio ({ratio_label})"] = _ratio_value(r, "qps_a", "qps_b")
            out["p99_a"] = r.get("p99_a", "")
            out["p99_b"] = r.get("p99_b", "")
            out[f"p99_ratio ({ratio_label})"] = _ratio_value(r, "p99_a", "p99_b")
            ws_compare.append([out.get(k, "") for k in compare_fields])

        ws_info = wb.create_sheet("server-info")
        ws_info.append(["field", f"A ({run_a})", f"B ({run_b})"])
        ws_info.append(["query mode", meta_a.get("query_mode") or "", meta_b.get("query_mode") or ""])
        ws_info.append(["cpu bind", meta_a.get("cpu_bind") or "", meta_b.get("cpu_bind") or ""])
        ws_info.append(["cpu info", _multiline_kv(meta_a.get("lscpu_info")), _multiline_kv(meta_b.get("lscpu_info"))])
        ws_info.cell(row=4, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        ws_info.cell(row=4, column=3).alignment = Alignment(wrap_text=True, vertical="top")

        out = io.BytesIO()
        wb.save(out)
        data = out.getvalue()

        filename = f"compare.{algo.lower()}.{run_a}.vs.{run_b}.{ratio_mode}.xlsx"
        return Response(
            content=data,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @app.get("/api/run/{dataset}/{run_id}/summary")
    def api_run_summary(dataset: str, run_id: str) -> JSONResponse:
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        headers, rows = _read_tsv(run_dir / "outputs" / "summary.tsv")
        return JSONResponse({"headers": headers, "rows": rows})

    @app.get("/api/run/{dataset}/{run_id}/mode")
    def api_run_mode(dataset: str, run_id: str) -> PlainTextResponse:
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        mode_txt = _read_text_if_exists(run_dir / "mode.txt", max_bytes=2000) or ""
        return PlainTextResponse(mode_txt)

    return app


def _default_runs_dir() -> Path:
    # Default to DiskANN-playground/diskann-ann-bench/result relative to this file.
    here = Path(__file__).resolve()
    # .../DiskANN-playground/diskann-ann-bench/web/app.py
    playground = here.parents[2]
    return (playground / "diskann-ann-bench" / "result").resolve()


def main() -> int:
    ap = argparse.ArgumentParser(description="diskann-ann-bench web")
    ap.add_argument("--host", default=_env("WEB_HOST", "127.0.0.1"))
    ap.add_argument("--port", type=int, default=int(_env("WEB_PORT", "8081")))
    ap.add_argument("--runs-dir", default=_env("RUNS_DIR", str(_default_runs_dir())))
    args = ap.parse_args()

    settings = Settings(host=args.host, port=int(args.port), runs_dir=Path(args.runs_dir).expanduser().resolve())

    app = create_app(settings)

    import uvicorn  # type: ignore

    uvicorn.run(app, host=settings.host, port=settings.port, log_level="info")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
