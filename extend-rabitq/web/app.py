#!/usr/bin/env python3

from __future__ import annotations

import argparse
import csv
import html
import io
import json
import os
import urllib.parse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from fastapi import FastAPI, HTTPException
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, PlainTextResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request

from markdown_it import MarkdownIt


@dataclass(frozen=True)
class Settings:
    host: str
    port: int
    runs_dir: Path


def _default_runs_dir() -> Path:
    here = Path(__file__).resolve()
    extend_rabitq_dir = here.parent.parent
    return (extend_rabitq_dir / "ann-harness" / "runs").resolve()


def _safe_join(root: Path, *parts: str) -> Path:
    out = root
    for p in parts:
        # Prevent absolute paths.
        if p.startswith("/") or p.startswith("\\"):
            raise HTTPException(status_code=400, detail="invalid path")
        out = out / p
    out = out.resolve()
    try:
        out.relative_to(root)
    except Exception:
        raise HTTPException(status_code=400, detail="path traversal blocked")
    return out


def _list_dirs(p: Path) -> list[str]:
    if not p.exists():
        return []
    out: list[str] = []
    for child in p.iterdir():
        if child.is_dir():
            out.append(child.name)
    out.sort()
    return out


def _list_dirs_sorted_desc(p: Path) -> list[str]:
    if not p.exists():
        return []
    dirs = [c for c in p.iterdir() if c.is_dir()]
    dirs.sort(key=lambda x: x.name, reverse=True)
    return [d.name for d in dirs]


def _read_text_if_exists(path: Path, *, max_bytes: int = 2_000_000) -> str | None:
    try:
        if not path.is_file():
            return None
        data = path.read_bytes()
        if len(data) > max_bytes:
            return f"<file too large: {len(data)} bytes>"
        return data.decode("utf-8", errors="replace")
    except Exception:
        return None


def _read_tsv(path: Path, *, max_rows: int = 5000) -> tuple[list[str], list[dict[str, str]]]:
    if not path.is_file():
        return ([], [])
    with path.open("r", encoding="utf-8", errors="replace", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        headers = reader.fieldnames or []
        rows: list[dict[str, str]] = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                rows.append({h: f"<truncated after {max_rows} rows>" for h in headers})
                break
            rows.append({k: (v if v is not None else "") for k, v in row.items()})
        return (headers, rows)


def _markdown_to_html(md_text: str) -> str:
    # Keep it simple and safe-ish:
    # - markdown-it-py does not execute JS.
    # - We still escape raw HTML by disabling html.
    md = MarkdownIt("commonmark", {"html": False, "linkify": True, "typographer": False})
    rendered = md.render(md_text)
    return rendered


def _extract_cpu_bind_from_server_info(server_info: str | None) -> str | None:
    if not server_info:
        return None
    for line in server_info.splitlines():
        # run_dataset.py writes: cpu_bind: <desc>
        if line.startswith("cpu_bind:"):
            v = line.split(":", 1)[1].strip()
            return v or None
    return None


def _extract_cpu_model_from_server_info(server_info: str | None) -> str | None:
    if not server_info:
        return None
    for line in server_info.splitlines():
        # lscpu line format on Linux:
        #   Model name:                           Intel(R) Xeon(R) ...
        if line.lower().startswith("model name:"):
            v = line.split(":", 1)[1].strip()
            return v or None
    return None


def _extract_numactl_cpulist(cpu_bind: str | None) -> str | None:
    if not cpu_bind:
        return None
    # Expect patterns like: "numactl -C 0-15 -m 0" (tokens) or "numactl -C0-15 -m0".
    tokens = cpu_bind.split()
    for i, tok in enumerate(tokens):
        if tok == "-C" and i + 1 < len(tokens):
            return tokens[i + 1].strip() or None
        if tok.startswith("-C") and len(tok) > 2:
            return tok[2:].strip() or None
        if tok.startswith("--physcpubind="):
            v = tok.split("=", 1)[1].strip()
            return v or None
    return None


def _count_cpulist(cpulist: str | None) -> int | None:
    if not cpulist:
        return None
    total = 0
    for part in cpulist.split(","):
        p = part.strip()
        if not p:
            continue
        if "-" in p:
            a, b, *rest = p.split("-")
            if rest:
                return None
            try:
                start = int(a)
                end = int(b)
            except ValueError:
                return None
            if end < start:
                return None
            total += (end - start + 1)
        else:
            try:
                int(p)
            except ValueError:
                return None
            total += 1
    return total


def _parse_sort_spec(spec: str | None) -> list[tuple[str, bool]]:
    # Returns list[(field, reverse)]
    # Supported fields: timestamp, cpu_cores, cpu_model
    default = [
        ("timestamp", True),
        ("cpu_cores", True),
        ("cpu_model", False),
    ]
    if not spec:
        return default

    allowed = {"timestamp", "cpu_cores", "cpu_model"}
    out: list[tuple[str, bool]] = []
    for raw in str(spec).split(","):
        item = raw.strip()
        if not item:
            continue
        reverse = item.startswith("-")
        field = item[1:] if reverse else item
        if field not in allowed:
            continue
        out.append((field, reverse))
    return out or default


def _to_int_maybe(v: str | None) -> int | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def _to_float_maybe(v: str | None) -> float | None:
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.lower() == "none":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _fmt_delta(v: float | None) -> str:
    if v is None:
        return "-"
    # Keep small deltas readable.
    return f"{v:+.4g}"


def _fmt_ratio_pct(numer: float | None, denom: float | None) -> str:
    # Return percent change: ((numer/denom) - 1) * 100%.
    # Example: +3.1% means numer is 3.1% higher than denom.
    if numer is None or denom is None:
        return "-"
    if denom == 0:
        return "-"
    return f"{((numer / denom) - 1.0) * 100.0:+.2f}%"


def _run_meta(run_dir: Path) -> dict[str, Any]:
    server_info_head = _read_text_if_exists(run_dir / "server-info.txt", max_bytes=300_000)
    cpu_bind = _read_text_if_exists(run_dir / "cpu-bind.txt")
    if not cpu_bind:
        cpu_bind = _extract_cpu_bind_from_server_info(server_info_head)

    cpu_model = _extract_cpu_model_from_server_info(server_info_head)
    cpu_cores = _count_cpulist(_extract_numactl_cpulist(cpu_bind))

    return {
        "cpu_bind": (cpu_bind.strip() if isinstance(cpu_bind, str) and cpu_bind.strip() else None),
        "cpu_model": cpu_model,
        "cpu_cores": cpu_cores,
        "server_info": server_info_head,
        "has_cpu_bind_file": (run_dir / "cpu-bind.txt").is_file(),
        "run_path": str(run_dir),
    }


def _index_rows(rows: list[dict[str, str]], key_fields: list[str]) -> dict[tuple[str, ...], dict[str, str]]:
    out: dict[tuple[str, ...], dict[str, str]] = {}
    for r in rows:
        k = tuple((r.get(f) or "").strip() for f in key_fields)
        out[k] = r
    return out


def _list_config_json_files(run_dir: Path) -> list[str]:
    cfg_dir = run_dir / "configs"
    if not cfg_dir.is_dir():
        return []
    out: list[str] = []
    for child in cfg_dir.iterdir():
        if child.is_file() and child.suffix.lower() == ".json":
            out.append(child.name)
    out.sort()
    return out


def _read_json_if_exists(path: Path, *, max_bytes: int = 2_000_000) -> Any | None:
    text = _read_text_if_exists(path, max_bytes=max_bytes)
    if text is None:
        return None
    try:
        return json.loads(text)
    except Exception:
        return None


def _flatten_json(
    obj: Any,
    *,
    max_items: int = 5000,
) -> tuple[dict[str, str], bool]:
    # Returns (flat_map, truncated)
    out: dict[str, str] = {}
    truncated = False

    def add(path: str, value: Any):
        nonlocal truncated
        if truncated:
            return
        if len(out) >= max_items:
            truncated = True
            return
        if value is None or isinstance(value, (str, int, float, bool)):
            out[path] = "" if value is None else str(value)
        else:
            # Fallback: stable JSON representation for non-primitive leaf.
            try:
                out[path] = json.dumps(value, ensure_ascii=False, sort_keys=True)
            except Exception:
                out[path] = str(value)

    def walk(current: Any, prefix: str):
        nonlocal truncated
        if truncated:
            return
        if current is None or isinstance(current, (str, int, float, bool)):
            add(prefix or "$", current)
            return
        if isinstance(current, dict):
            if not current:
                add(prefix or "$", {})
                return
            for k in sorted(current.keys(), key=lambda x: str(x)):
                v = current.get(k)
                key = str(k)
                next_prefix = f"{prefix}.{key}" if prefix else key
                walk(v, next_prefix)
            return
        if isinstance(current, list):
            if not current:
                add(prefix or "$", [])
                return
            for i, v in enumerate(current):
                next_prefix = f"{prefix}[{i}]" if prefix else f"[{i}]"
                walk(v, next_prefix)
            return

        add(prefix or "$", current)

    walk(obj, "")
    return (out, truncated)


def _diff_flat_maps(
    a_map: dict[str, str],
    b_map: dict[str, str],
    *,
    max_rows: int = 200,
) -> tuple[list[dict[str, str]], int, bool]:
    keys = sorted(set(a_map.keys()) | set(b_map.keys()))
    rows: list[dict[str, str]] = []
    diff_count = 0
    truncated = False
    for k in keys:
        av = a_map.get(k, "")
        bv = b_map.get(k, "")
        if av == bv:
            continue
        diff_count += 1
        if len(rows) >= max_rows:
            truncated = True
            continue
        rows.append({"path": k, "a": av, "b": bv})
    return (rows, diff_count, truncated)


def _truncate_str(s: str, *, max_len: int = 200) -> str:
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def _diff_flat_maps_rich(
    a_map: dict[str, str],
    b_map: dict[str, str],
    *,
    max_rows: int = 200,
    max_value_len: int = 200,
) -> tuple[list[dict[str, str]], int, bool]:
    keys = sorted(set(a_map.keys()) | set(b_map.keys()))
    rows: list[dict[str, str]] = []
    diff_count = 0
    truncated = False
    for k in keys:
        a_present = k in a_map
        b_present = k in b_map
        av_full = a_map.get(k, "")
        bv_full = b_map.get(k, "")
        if a_present and b_present and av_full == bv_full:
            continue
        if not a_present and not b_present:
            continue

        if not a_present and b_present:
            status = "added"
        elif a_present and not b_present:
            status = "removed"
        else:
            status = "changed"

        diff_count += 1
        if len(rows) >= max_rows:
            truncated = True
            continue
        rows.append(
            {
                "path": k,
                "status": status,
                "a": _truncate_str(av_full, max_len=max_value_len),
                "b": _truncate_str(bv_full, max_len=max_value_len),
                "a_full": av_full,
                "b_full": bv_full,
            }
        )
    return (rows, diff_count, truncated)


def _parse_lscpu_kv(server_info: str | None) -> dict[str, str]:
    if not server_info:
        return {}
    lines = server_info.splitlines()
    try:
        i = lines.index("lscpu:")
    except ValueError:
        return {}
    out: dict[str, str] = {}
    for line in lines[i + 1 :]:
        if not line.strip():
            continue
        # lscpu output is generally "Key:   Value".
        if ":" not in line:
            continue
        k, v = line.split(":", 1)
        k = k.strip()
        v = v.strip()
        if not k:
            continue
        if k not in out:
            out[k] = v
    return out


def _parse_details_job_configs(details_md: str | None) -> dict[str, dict[str, str]]:
    # Parse outputs/details.md and extract shallow (non-nested) job config keys.
    if not details_md:
        return {}
    jobs: dict[str, dict[str, str]] = {}
    current_job: str | None = None
    for raw in details_md.splitlines():
        line = raw.rstrip("\n")
        if line.startswith("## Job ") and ":" in line:
            # Example: "## Job 1: async-index-build-pq"
            current_job = line.split(":", 1)[1].strip() or None
            if current_job and current_job not in jobs:
                jobs[current_job] = {}
            continue
        if not current_job:
            continue
        # Only shallow bullets (no leading spaces)
        if line.startswith("- ") and not line.startswith("  "):
            # Example: "- pq_chunks: 64"; ignore nested section markers like "- build:".
            body = line[2:]
            if ":" not in body:
                continue
            k, v = body.split(":", 1)
            k = k.strip()
            v = v.strip()
            if not k:
                continue
            if not v:
                continue
            jobs[current_job][k] = v
    return jobs


def _try_build_xlsx_compare(
    *,
    dataset: str,
    a: str,
    b: str,
    mode: str,
    delta_label: str,
    meta_a: dict[str, Any],
    meta_b: dict[str, Any],
    key_fields: list[str],
    metrics: list[str],
    rows: list[dict[str, Any]],
    run_a_dir: Path,
    run_b_dir: Path,
) -> bytes:
    try:
        from openpyxl import Workbook  # type: ignore[import-not-found]
        from openpyxl.styles import Alignment, Font  # type: ignore[import-not-found]
    except Exception as e:  # noqa: BLE001
        raise RuntimeError(
            "Missing dependency 'openpyxl'. Install with: conda install -n diskann-rs -c conda-forge openpyxl"
        ) from e

    wb = Workbook()

    # Sheet 1: CPU info
    ws = wb.active
    ws.title = "CPUInfo"
    ws.append(["field", "A", "B"])
    ws["A1"].font = ws["B1"].font = ws["C1"].font = Font(bold=True)
    ws.freeze_panes = "A2"

    def add(field: str, va: Any, vb: Any):
        ws.append([field, "" if va is None else str(va), "" if vb is None else str(vb)])

    add("dataset", dataset, dataset)
    add("run_id", a, b)
    add("run_path", meta_a.get("run_path"), meta_b.get("run_path"))
    add("cpu-bind.txt", "yes" if meta_a.get("has_cpu_bind_file") else "no", "yes" if meta_b.get("has_cpu_bind_file") else "no")
    add("cpu_cores", meta_a.get("cpu_cores"), meta_b.get("cpu_cores"))
    add("cpu_model", meta_a.get("cpu_model"), meta_b.get("cpu_model"))
    add("cpu_bind", meta_a.get("cpu_bind"), meta_b.get("cpu_bind"))

    lscpu_a = _parse_lscpu_kv(meta_a.get("server_info"))
    lscpu_b = _parse_lscpu_kv(meta_b.get("server_info"))
    important = [
        "CPU(s)",
        "On-line CPU(s) list",
        "Thread(s) per core",
        "Core(s) per socket",
        "Socket(s)",
        "NUMA node(s)",
        "NUMA node0 CPU(s)",
        "NUMA node1 CPU(s)",
    ]
    for k in important:
        if k in lscpu_a or k in lscpu_b:
            add(f"lscpu.{k}", lscpu_a.get(k), lscpu_b.get(k))

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 60

    # Sheet 2: summary.tsv compare (flat, wide)
    ws2 = wb.create_sheet("SummaryCompare")
    ws2.append(["dataset", dataset])
    ws2.append(["A", a, "B", b, "mode", mode, "delta", delta_label])
    ws2.append([])
    header = list(key_fields)
    for m in metrics:
        header.extend([f"{m} (A)", f"{m} (B)", f"{m} Δ%"])
    ws2.append(header)
    for cell in ws2[4]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws2.freeze_panes = "A5"

    for r in rows:
        out_row: list[Any] = []
        for k in key_fields:
            out_row.append(r.get(k, ""))
        metric_list = r.get("metrics") or []
        by_name = {m["name"]: m for m in metric_list if isinstance(m, dict) and "name" in m}
        for m in metrics:
            mr = by_name.get(m, {})
            out_row.extend([mr.get("a", "-"), mr.get("b", "-"), mr.get("ratio_pct", "-")])
        ws2.append(out_row)

    # Sheet 3: details.md job config compare
    ws3 = wb.create_sheet("JobConfigCompare")
    ws3.append(["job", "key", "A", "B", "same"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    ws3.freeze_panes = "A2"

    details_a = _read_text_if_exists(run_a_dir / "outputs" / "details.md", max_bytes=2_000_000)
    details_b = _read_text_if_exists(run_b_dir / "outputs" / "details.md", max_bytes=2_000_000)
    jobs_a = _parse_details_job_configs(details_a)
    jobs_b = _parse_details_job_configs(details_b)
    all_jobs = sorted(set(jobs_a.keys()) | set(jobs_b.keys()))
    for job in all_jobs:
        ka = jobs_a.get(job, {})
        kb = jobs_b.get(job, {})
        keys = sorted(set(ka.keys()) | set(kb.keys()))
        if not keys:
            ws3.append([job, "<no shallow config parsed>", "", "", ""])
            continue
        for k in keys:
            va = ka.get(k, "")
            vb = kb.get(k, "")
            ws3.append([job, k, va, vb, "yes" if (va == vb and va != "") else ("no" if (va or vb) else "")])

    # Wrap text for long cells.
    for w in (ws, ws2, ws3):
        for row in w.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _breadcrumbs(*items: tuple[str, str]) -> str:
    # items: (label, href)
    parts: list[str] = []
    for label, href in items:
        parts.append(f'<a href="{html.escape(href, quote=True)}">{html.escape(label)}</a>')
    return " / ".join(parts)


def create_app(settings: Settings) -> FastAPI:
    app = FastAPI(title="extend-rabitq web")

    templates_dir = Path(__file__).resolve().parent / "templates"
    templates = Jinja2Templates(directory=str(templates_dir))
    templates.env.filters["urlencode"] = lambda s: urllib.parse.quote(str(s), safe="")

    static_dir = Path(__file__).resolve().parent / "static"
    app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

    @app.get("/api/health")
    def health() -> dict[str, Any]:
        return {
            "ok": True,
            "runs_dir": str(settings.runs_dir),
        }

    @app.get("/api/datasets")
    def api_datasets() -> JSONResponse:
        return JSONResponse(_list_dirs(settings.runs_dir))

    @app.get("/api/runs/{dataset}")
    def api_runs(dataset: str) -> JSONResponse:
        dataset_dir = _safe_join(settings.runs_dir, dataset)
        return JSONResponse(_list_dirs_sorted_desc(dataset_dir))

    @app.get("/api/run/{dataset}/{run_id}/summary")
    def api_run_summary(dataset: str, run_id: str) -> JSONResponse:
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        summary_path = run_dir / "outputs" / "summary.tsv"
        headers, rows = _read_tsv(summary_path)
        return JSONResponse({"headers": headers, "rows": rows})

    @app.get("/", response_class=HTMLResponse)
    def index(request: Request):
        datasets = _list_dirs(settings.runs_dir)
        return templates.TemplateResponse(
            "index.html",
            {
                "request": request,
                "title": "DiskANN runs",
                "datasets": datasets,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": None,
            },
        )

    @app.get("/dataset/{dataset}", response_class=HTMLResponse)
    def dataset_page(dataset: str, request: Request):
        dataset_dir = _safe_join(settings.runs_dir, dataset)
        run_ids = _list_dirs_sorted_desc(dataset_dir)
        sort_spec_raw = request.query_params.get("sort")
        sort_spec = _parse_sort_spec(sort_spec_raw)
        runs = []
        for run_id in run_ids:
            run_dir = dataset_dir / run_id
            server_info_head = _read_text_if_exists(run_dir / "server-info.txt", max_bytes=200_000)
            cpu_bind = _read_text_if_exists(run_dir / "cpu-bind.txt")
            if not cpu_bind:
                cpu_bind = _extract_cpu_bind_from_server_info(server_info_head)

            cpu_model = _extract_cpu_model_from_server_info(server_info_head)
            cpu_cores = _count_cpulist(_extract_numactl_cpulist(cpu_bind))

            runs.append(
                {
                    "id": run_id,
                    "timestamp": run_id,
                    "has_cpu_bind": (run_dir / "cpu-bind.txt").is_file(),
                    "cpu_bind": (cpu_bind.strip() if isinstance(cpu_bind, str) and cpu_bind.strip() else None),
                    "cpu_cores": cpu_cores,
                    "cpu_model": cpu_model,
                }
            )

        # Apply multi-key sort. We sort last key first (stable sort) to honor sort_spec order.
        for field, reverse in reversed(sort_spec):
            if field == "timestamp":
                runs.sort(key=lambda r: r.get("timestamp") or "", reverse=reverse)
            elif field == "cpu_cores":
                if reverse:
                    runs.sort(
                        key=lambda r: (
                            r.get("cpu_cores") is None,
                            -(int(r.get("cpu_cores")) if r.get("cpu_cores") is not None else 0),
                        )
                    )
                else:
                    runs.sort(
                        key=lambda r: (
                            r.get("cpu_cores") is None,
                            int(r.get("cpu_cores")) if r.get("cpu_cores") is not None else 0,
                        )
                    )
            elif field == "cpu_model":
                runs.sort(key=lambda r: (r.get("cpu_model") or ""), reverse=reverse)
                # Keep unknown CPU model (None) last regardless of direction.
                runs.sort(key=lambda r: (r.get("cpu_model") is None))
        bc = _breadcrumbs(("datasets", "/"), (dataset, f"/dataset/{dataset}"))
        return templates.TemplateResponse(
            "dataset.html",
            {
                "request": request,
                "title": f"{dataset} runs",
                "dataset": dataset,
                "runs": runs,
                "sort": sort_spec_raw or "-timestamp,-cpu_cores,cpu_model",
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": bc,
            },
        )

    @app.get("/run/{dataset}/{run_id}", response_class=HTMLResponse)
    def run_page(dataset: str, run_id: str, request: Request):
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        if not run_dir.is_dir():
            raise HTTPException(status_code=404, detail="run not found")

        cpu_bind_path = run_dir / "cpu-bind.txt"
        has_cpu_bind_file = cpu_bind_path.is_file()

        cpu_bind = _read_text_if_exists(cpu_bind_path)

        summary_path = run_dir / "outputs" / "summary.tsv"
        headers, rows = _read_tsv(summary_path)

        details_text = _read_text_if_exists(run_dir / "outputs" / "details.md")
        details_html = _markdown_to_html(details_text) if details_text else None

        server_info = _read_text_if_exists(run_dir / "server-info.txt")
        if not cpu_bind:
            cpu_bind = _extract_cpu_bind_from_server_info(server_info)

        downloads = []
        for rel, label in [
            ("cpu-bind.txt", "cpu-bind.txt"),
            ("outputs/summary.tsv", "summary.tsv"),
            ("outputs/details.md", "details.md"),
            ("outputs/output.json", "output.json"),
            ("outputs/output.search.json", "output.search.json"),
            ("outputs/output.build.json", "output.build.json"),
            ("configs/pq-vs-spherical.json", "config"),
            ("server-info.txt", "server-info.txt"),
        ]:
            p = run_dir / rel
            if p.is_file():
                downloads.append({
                    "label": label,
                    "href": f"/file/{dataset}/{run_id}/{rel}",
                })

        bc = _breadcrumbs(
            ("datasets", "/"),
            (dataset, f"/dataset/{dataset}"),
            (run_id, f"/run/{dataset}/{run_id}"),
        )
        return templates.TemplateResponse(
            "run.html",
            {
                "request": request,
                "title": f"{dataset}/{run_id}",
                "dataset": dataset,
                "run_id": run_id,
                "run_path": str(run_dir),
                "has_cpu_bind_file": has_cpu_bind_file,
                "cpu_bind": (cpu_bind.strip() if isinstance(cpu_bind, str) and cpu_bind.strip() else None),
                "summary_headers": headers,
                "summary_rows": rows,
                "details_html": details_html,
                "server_info": server_info,
                "downloads": downloads,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": bc,
            },
        )

    @app.get("/compare/{dataset}", response_class=HTMLResponse)
    def compare_page(dataset: str, request: Request):
        a = request.query_params.get("a")
        b = request.query_params.get("b")
        mode = (request.query_params.get("mode") or "b_over_a").strip().lower()
        if mode in {"ba", "b/a", "b_over_a"}:
            mode = "b_over_a"
        elif mode in {"ab", "a/b", "a_over_b"}:
            mode = "a_over_b"
        else:
            mode = "b_over_a"

        delta_label = "Δ = ((B/A) - 1) × 100%" if mode == "b_over_a" else "Δ = ((A/B) - 1) × 100%"
        if not a or not b:
            raise HTTPException(status_code=400, detail="missing query params: a, b")
        if a == b:
            raise HTTPException(status_code=400, detail="a and b must be different")

        dataset_dir = _safe_join(settings.runs_dir, dataset)
        run_a_dir = _safe_join(dataset_dir, a)
        run_b_dir = _safe_join(dataset_dir, b)
        if not run_a_dir.is_dir() or not run_b_dir.is_dir():
            raise HTTPException(status_code=404, detail="run not found")

        meta_a = _run_meta(run_a_dir)
        meta_b = _run_meta(run_b_dir)

        # configs/*.json compare (optional)
        cfg_names_a = set(_list_config_json_files(run_a_dir))
        cfg_names_b = set(_list_config_json_files(run_b_dir))
        cfg_names = sorted(cfg_names_a | cfg_names_b)
        config_compares: list[dict[str, Any]] = []
        for name in cfg_names:
            pa = (run_a_dir / "configs" / name)
            pb = (run_b_dir / "configs" / name)
            present_a = pa.is_file()
            present_b = pb.is_file()
            raw_a = _read_text_if_exists(pa, max_bytes=400_000) if present_a else None
            raw_b = _read_text_if_exists(pb, max_bytes=400_000) if present_b else None

            ja = _read_json_if_exists(pa) if present_a else None
            jb = _read_json_if_exists(pb) if present_b else None
            parse_a_ok = ja is not None
            parse_b_ok = jb is not None

            diffs: list[dict[str, str]] = []
            diff_count = 0
            diff_truncated = False
            flatten_truncated = False
            if parse_a_ok and parse_b_ok:
                flat_a, trunc_a = _flatten_json(ja)
                flat_b, trunc_b = _flatten_json(jb)
                flatten_truncated = trunc_a or trunc_b
                diffs, diff_count, diff_truncated = _diff_flat_maps_rich(flat_a, flat_b)

            config_compares.append(
                {
                    "name": name,
                    "present_a": present_a,
                    "present_b": present_b,
                    "a_href": f"/file/{urllib.parse.quote(dataset)}/{urllib.parse.quote(a)}/configs/{urllib.parse.quote(name)}" if present_a else None,
                    "b_href": f"/file/{urllib.parse.quote(dataset)}/{urllib.parse.quote(b)}/configs/{urllib.parse.quote(name)}" if present_b else None,
                    "parse_a_ok": parse_a_ok,
                    "parse_b_ok": parse_b_ok,
                    "diffs": diffs,
                    "diff_count": diff_count,
                    "diff_truncated": diff_truncated,
                    "flatten_truncated": flatten_truncated,
                    "raw_a": raw_a,
                    "raw_b": raw_b,
                }
            )

        headers_a, rows_a = _read_tsv(run_a_dir / "outputs" / "summary.tsv")
        headers_b, rows_b = _read_tsv(run_b_dir / "outputs" / "summary.tsv")

        # Compare only on common columns.
        common_headers = [h for h in headers_a if h in set(headers_b)]
        key_fields = [f for f in ["job", "detail", "tasks", "L", "N"] if f in common_headers]
        metric_fields = [h for h in common_headers if h not in key_fields]

        # Show ALL metrics; order a few important ones first.
        preferred_metrics = [
            "recall(avg)",
            "QPS(mean)",
            "lat_mean_us(mean)",
            "lat_p99_us(mean)",
        ]
        metrics = [m for m in preferred_metrics if m in metric_fields]
        metrics.extend([m for m in metric_fields if m not in set(metrics)])

        idx_a = _index_rows(rows_a, key_fields)
        idx_b = _index_rows(rows_b, key_fields)
        all_keys = sorted(set(idx_a.keys()) | set(idx_b.keys()))

        def sort_key(k: tuple[str, ...]) -> tuple[Any, ...]:
            # job/detail lexicographic; tasks/L/N numeric when possible.
            parts: list[Any] = []
            for name, v in zip(key_fields, k):
                if name in {"tasks", "L", "N"}:
                    parts.append(_to_int_maybe(v) if _to_int_maybe(v) is not None else v)
                else:
                    parts.append(v)
            return tuple(parts)

        all_keys.sort(key=sort_key)

        compare_rows: list[dict[str, Any]] = []
        for k in all_keys:
            ra = idx_a.get(k)
            rb = idx_b.get(k)
            out: dict[str, Any] = {}
            for i, f in enumerate(key_fields):
                out[f] = k[i] if i < len(k) else ""
            out["present_a"] = ra is not None
            out["present_b"] = rb is not None

            metric_rows: list[dict[str, Any]] = []
            for m in metrics:
                va = (ra.get(m) if ra else None)
                vb = (rb.get(m) if rb else None)
                da = _to_float_maybe(va)
                db = _to_float_maybe(vb)
                delta = _fmt_ratio_pct(db, da) if mode == "b_over_a" else _fmt_ratio_pct(da, db)
                metric_rows.append(
                    {
                        "name": m,
                        "a": (va if va is not None and str(va).strip() else "-"),
                        "b": (vb if vb is not None and str(vb).strip() else "-"),
                        "ratio_pct": delta,
                    }
                )
            out["metrics"] = metric_rows
            compare_rows.append(out)

        bc = _breadcrumbs(
            ("datasets", "/"),
            (dataset, f"/dataset/{dataset}"),
            ("compare", f"/compare/{dataset}?a={urllib.parse.quote(a)}&b={urllib.parse.quote(b)}&mode={urllib.parse.quote(mode)}"),
        )
        return templates.TemplateResponse(
            "compare.html",
            {
                "request": request,
                "title": f"compare {dataset}",
                "dataset": dataset,
                "a": a,
                "b": b,
                "meta_a": meta_a,
                "meta_b": meta_b,
                "mode": mode,
                "delta_label": delta_label,
                "download_href": f"/compare/{dataset}/download?a={urllib.parse.quote(a)}&b={urllib.parse.quote(b)}&mode={urllib.parse.quote(mode)}",
                "config_compares": config_compares,
                "key_fields": key_fields,
                "metrics": metrics,
                "rows": compare_rows,
                "wide": True,
                "runs_dir": str(settings.runs_dir),
                "breadcrumbs": bc,
            },
        )

    @app.get("/compare/{dataset}/download")
    def compare_download(dataset: str, request: Request):
        a = request.query_params.get("a")
        b = request.query_params.get("b")
        mode = (request.query_params.get("mode") or "b_over_a").strip().lower()
        if mode in {"ba", "b/a", "b_over_a"}:
            mode = "b_over_a"
        elif mode in {"ab", "a/b", "a_over_b"}:
            mode = "a_over_b"
        else:
            mode = "b_over_a"

        delta_label = "Δ = ((B/A) - 1) × 100%" if mode == "b_over_a" else "Δ = ((A/B) - 1) × 100%"
        if not a or not b:
            raise HTTPException(status_code=400, detail="missing query params: a, b")
        if a == b:
            raise HTTPException(status_code=400, detail="a and b must be different")

        dataset_dir = _safe_join(settings.runs_dir, dataset)
        run_a_dir = _safe_join(dataset_dir, a)
        run_b_dir = _safe_join(dataset_dir, b)
        if not run_a_dir.is_dir() or not run_b_dir.is_dir():
            raise HTTPException(status_code=404, detail="run not found")

        meta_a = _run_meta(run_a_dir)
        meta_b = _run_meta(run_b_dir)

        headers_a, rows_a = _read_tsv(run_a_dir / "outputs" / "summary.tsv")
        headers_b, rows_b = _read_tsv(run_b_dir / "outputs" / "summary.tsv")
        common_headers = [h for h in headers_a if h in set(headers_b)]
        key_fields = [f for f in ["job", "detail", "tasks", "L", "N"] if f in common_headers]
        metric_fields = [h for h in common_headers if h not in key_fields]
        preferred_metrics = [
            "recall(avg)",
            "QPS(mean)",
            "lat_mean_us(mean)",
            "lat_p99_us(mean)",
        ]
        metrics = [m for m in preferred_metrics if m in metric_fields]
        metrics.extend([m for m in metric_fields if m not in set(metrics)])

        idx_a = _index_rows(rows_a, key_fields)
        idx_b = _index_rows(rows_b, key_fields)
        all_keys = sorted(set(idx_a.keys()) | set(idx_b.keys()))

        def sort_key(k: tuple[str, ...]) -> tuple[Any, ...]:
            parts: list[Any] = []
            for name, v in zip(key_fields, k):
                if name in {"tasks", "L", "N"}:
                    parts.append(_to_int_maybe(v) if _to_int_maybe(v) is not None else v)
                else:
                    parts.append(v)
            return tuple(parts)

        all_keys.sort(key=sort_key)

        compare_rows: list[dict[str, Any]] = []
        for k in all_keys:
            ra = idx_a.get(k)
            rb = idx_b.get(k)
            out: dict[str, Any] = {}
            for i, f in enumerate(key_fields):
                out[f] = k[i] if i < len(k) else ""
            out["present_a"] = ra is not None
            out["present_b"] = rb is not None

            metric_rows: list[dict[str, Any]] = []
            for m in metrics:
                va = (ra.get(m) if ra else None)
                vb = (rb.get(m) if rb else None)
                da = _to_float_maybe(va)
                db = _to_float_maybe(vb)
                delta = _fmt_ratio_pct(db, da) if mode == "b_over_a" else _fmt_ratio_pct(da, db)
                metric_rows.append(
                    {
                        "name": m,
                        "a": (va if va is not None and str(va).strip() else "-"),
                        "b": (vb if vb is not None and str(vb).strip() else "-"),
                        "ratio_pct": delta,
                    }
                )
            out["metrics"] = metric_rows
            compare_rows.append(out)

        xlsx_bytes = _try_build_xlsx_compare(
            dataset=dataset,
            a=a,
            b=b,
            mode=mode,
            delta_label=delta_label,
            meta_a=meta_a,
            meta_b=meta_b,
            key_fields=key_fields,
            metrics=metrics,
            rows=compare_rows,
            run_a_dir=run_a_dir,
            run_b_dir=run_b_dir,
        )

        filename = f"compare_{dataset}_{a}_vs_{b}_{mode}.xlsx"
        headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
        return StreamingResponse(io.BytesIO(xlsx_bytes), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)

    @app.get("/file/{dataset}/{run_id}/{rel_path:path}")
    def get_file(dataset: str, run_id: str, rel_path: str):
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        file_path = _safe_join(run_dir, rel_path)
        if not file_path.is_file():
            raise HTTPException(status_code=404, detail="file not found")
        return FileResponse(str(file_path), filename=file_path.name)

    @app.get("/raw/{dataset}/{run_id}/{rel_path:path}")
    def get_raw(dataset: str, run_id: str, rel_path: str):
        run_dir = _safe_join(settings.runs_dir, dataset, run_id)
        file_path = _safe_join(run_dir, rel_path)
        if not file_path.is_file():
            raise HTTPException(status_code=404, detail="file not found")
        text = _read_text_if_exists(file_path)
        if text is None:
            raise HTTPException(status_code=500, detail="failed to read file")
        return PlainTextResponse(text)

    return app


def _parse_args() -> Settings:
    ap = argparse.ArgumentParser(description="Web UI to browse ann-harness runs")
    ap.add_argument("--host", default=os.environ.get("WEB_HOST", "127.0.0.1"))
    ap.add_argument("--port", type=int, default=int(os.environ.get("WEB_PORT", "8080")))
    ap.add_argument("--runs-dir", default=os.environ.get("RUNS_DIR", str(_default_runs_dir())))
    args = ap.parse_args()

    runs_dir = Path(args.runs_dir).expanduser().resolve()
    return Settings(host=str(args.host), port=int(args.port), runs_dir=runs_dir)


def main() -> int:
    settings = _parse_args()
    if not settings.runs_dir.exists():
        raise SystemExit(f"RUNS_DIR does not exist: {settings.runs_dir}")

    app = create_app(settings)

    import uvicorn  # type: ignore

    uvicorn.run(app, host=settings.host, port=settings.port, log_level="info")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
